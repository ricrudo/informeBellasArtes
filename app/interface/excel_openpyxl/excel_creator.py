from app.interface import db
from app.interface.db.get_Data import getTables
from app.domain import enums

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

import json
from pathlib import Path
from datetime import datetime

THIN = Side(border_style="thin", color="000000")
URL = 'https://riedmusicapp.com/informebellasartes/getsoportes/'

def get_informe_final():
    wb = Workbook()
    persons = get_persons()
    activesections = get_activesections()
    create_section1(wb, persons)
    create_section2(wb, persons)
    create_section3(wb, persons)
    create_section4_1(wb, persons, activesections)
    create_section4_2(wb, persons)
    create_section4_3(wb, persons, activesections)
    create_section4_4(wb, persons, activesections)
    create_section4_5(wb, persons, activesections)
    create_section4_6(wb, persons, activesections)
    create_section4_7(wb, persons, activesections)
    create_section4_8(wb, persons, activesections)
    create_section4_9(wb, persons, activesections)
    create_section4_10(wb, persons, activesections)
    create_section4_11(wb, persons, activesections)
    create_section4_12(wb, persons, activesections)
    create_section4_13(wb, persons, activesections)
    create_section4_14(wb, persons, activesections)
    create_section4_15(wb, persons, activesections)
    create_section4_16(wb, persons, activesections)
    create_section4_17(wb, persons, activesections)
    create_section5_1(wb, persons, activesections)
    create_section5_2(wb, persons, activesections)
    create_section5_3(wb, persons, activesections)
    create_section6_1(wb, persons)
    create_section6_2(wb, persons, activesections)
    create_section6_3(wb, persons, activesections)
    create_section6_4(wb, persons, activesections)
    #create_section6_5(wb, persons, activesections)
    #create_section6_6(wb, persons, activesections)
    create_section6_7(wb, persons, activesections)
    create_activateSections(wb, persons, activesections)
    update_autofilter(wb)
    return savefile(wb)

def savefile(wb):
    name_file = f'informe_final_{datetime.utcnow().isoformat()}.xlsx'
    dirfile = Path.cwd() / 'userFiles' / 'informe_final'
    if not dirfile.exists():
        if not dirfile.parent.exists():
            dirfile.parent.mkdir()
        dirfile.mkdir()

    filename = dirfile / name_file
    wb.save(filename.as_posix())
    return filename.name


def get_persons():
    persons = getTables('section1_1')[0]
    data = db.session.query(persons).all()
    return {p.id: p.name for p in data}

def get_activesections():
    active_sections = getTables('activesections')[0]
    data = db.session.query(active_sections).all()
    response = {}
    for info in data:
        inneresponse = {}
        for attr in dir(info):
            if 'activateQuestionsection' in attr:
                inneresponse[attr.replace("activateQuestion","")] = getattr(info, attr)
        response[info.id_person] = inneresponse
    return response
   
def createTitles(ws, columnTitles, title, includeDocente=True):
    cell = ws.cell(row=1, column=1, value=title.upper())
    cell.font = Font(bold=True)
    if includeDocente:
        columnTitles = [('id Docente',13), ('docente',40)] + columnTitles
    row = 2
    for col, colName in enumerate(columnTitles, 1):
        cell = ws.cell(row=row, column=col, value=colName[0].upper())
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="0099CCFF")
        cell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
        ws.column_dimensions[cell.column_letter].width=colName[1]

def extractorToExcel(dataOrigin, ws, section, persons, fields, activesections=None, wrap_text=[], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=[], excludeofEmptyness=[], allowEmptyness=False, translateResponse=[]):
    '''
    forcedValueFields = [["namefield1","namefield2"],["valueforfield1","valueforfield2"]]
    valueFromEnum = [["namefield1","namefield2"],["enumClass1","enumClass2"]]
    specialCases = [["namefield1","namefield2"],["specialFunction1","specialFunction2"]]
    translateResponse = [["namefield1", "namefield1",],[{"original1_1": "new1_1", "original1_2": "new1_2"},{"original2_1": "new2_1", "original2_2": "new2_2"}]]
    '''
    cells = []
    if row is None:
        row = 2
    for entry in dataOrigin:
        if activesections and activesections[entry.id_person][section] != 'si':
            continue
        if not allowEmptyness and all([getattr(entry, namefield) is None for namefield in fields if namefield not in excludeofEmptyness]):
            continue
        row += 1
        cells.append(ws.cell(row=row, column=1, value=entry.id_person))
        cells.append(ws.cell(row=row, column=2, value=persons[entry.id_person]))
        for col, namefield in enumerate(fields, 3):
            if namefield == 'name_file':
                value = f'{URL}/{entry.period_spam}/{section}/{entry.name_file}' if entry.name_file else ''
            elif specialCases and namefield in specialCases[0]:
                value = specialCases[1][specialCases[0].index(namefield)](entry)
            elif forcedValueFields and namefield in forcedValueFields[0]:
                value = forcedValueFields[1][forcedValueFields[0].index(namefield)]
            elif namefield in itemsList:
                value = ", ".join(json.loads(getattr(entry, namefield))) if getattr(entry, namefield) else ''
            elif valueFromEnum and namefield in valueFromEnum[0]:
                enumClass = valueFromEnum[1][valueFromEnum[0].index(namefield)]
                value = enumClass[getattr(entry, namefield).replace("-", "_")].value if getattr(entry, namefield) else ''
            else:
                value = getattr(entry, namefield) if getattr(entry, namefield) else ''
                if translateResponse and namefield in translateResponse[0]:
                    value = translateResponse[1][value]

            cells.append(ws.cell(row=row, column=col, value=value))
            if namefield in wrap_text:
                cells[-1].alignment = Alignment(wrap_text=True)
            if namefield == 'name_file':
                cells[-1].hyperlink = value
    for cell in cells:
        cell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)

    return row

def update_autofilter(wb):
    for sheet in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(sheet)
        tableSize = ws.calculate_dimension().replace('A1', 'A2')
        ws.auto_filter.ref = tableSize


def create_section1(wb, persons):
    sectionName = 'section1_1'
    ws = wb.active
    ws.title = "Sección 1.1"
    title = "1. INFORMACIÓN DOCENTE"
    tables = getTables(sectionName)


    columnTitles = [
        ("id Docente", 13),
        ("Nombre", 40),
        ("email", 50),
        ("Nivel máximo formación", 30)
    ]

    createTitles(ws, columnTitles, title, includeDocente=False)

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()
    
    row = 2
    for entry in data:
        row += 1
        cells = []
        cells.append(ws.cell(row=row, column=1, value=entry.id))
        cells.append(ws.cell(row=row, column=2, value=entry.name))
        cells.append(ws.cell(row=row, column=3, value=entry.email))
        cells.append(ws.cell(row=row, column=4, value=entry.formacion))
        for cell in cells:
            cell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)


def create_section2(wb, persons):
    sectionName = 'section2_1'
    ws = wb.create_sheet("Sección 2")
    title = "2. Misional Docencia"

    columnTitles = [
        ("Nombre de la asignatura",40),
        ("Código",10),
        ("Programa",10),
        ("Realizado",10),
        ("Observaciones sobre el desarrollo de las actividades docentes",60),
        ("Actividades extracurriculares con componente internacional",60),
        ("Prácticas de innovación educativa con TIC",60),
        ("Motivo de cancelación",40)
    ]
    createTitles(ws, columnTitles, title)

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).filter(tables[0].status_done != None).all()
    fields = ['materia', 'code', 'programa', 'status_done', 'observaciones', 'internacional', 'tic', 'motivos_cancel']

    row = extractorToExcel(data, ws, sectionName, persons, fields, wrap_text=['observaciones', 'internacional', 'tic', 'motivos_cancel'])

    data = db.session.query(tables[1]).all()
    fields = ['materia_add', 'code_add', 'programa_add', 'status_done', 'observaciones_add', 'internacional_add', 'tic_add']

    forcedValueFields = [["status_done"],["si"]]

    extractorToExcel(data, ws, sectionName, persons, fields, wrap_text=['observaciones_add', 'internacional_add', 'tic_add'], forcedValueFields=forcedValueFields, row=row, excludeofEmptyness=["status_done"])

def specialSection3(entry):
    if entry.type_ == 'programa':
        return 'programa - ' + enums.ProgramasFacultad[getattr(entry, 'programa_')].value if getattr(entry, 'programa_') else 'programa - no especifica'
    elif entry.type_ == 'misional':
        return 'misional - ' + enums.MisionalesFacultad[getattr(entry, 'misional_')].value if getattr(entry, 'misional_') else 'misional - no especifica'
    else:
        return ''

def create_section3(wb, persons):
    sectionName = 'section3_1'
    ws = wb.create_sheet("Sección 3")
    title = "3. Participación procesos curriculares"
    columnTitles = [
        ("proceso", 40),
        ("instancia", 40),
        ("observaciones", 80)
    ]
    createTitles(ws, columnTitles, title)

    fields = ['proceso', 'type_', 'observaciones_']
    valueFromEnum = [['proceso'],[enums.PartProcesosNames]]
    specialCases = [['type_'],[specialSection3]]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).filter(tables[0].status_=='si').all()

    extractorToExcel(data, ws, sectionName, persons, fields, wrap_text=['observaciones_'], valueFromEnum=valueFromEnum, specialCases=specialCases)


def specialCase4_1(entry):
    lineas = json.loads(entry.linea_investiga)
    return ", ".join([enums.LineasInvestigacionFacultad[k].value for k,v in lineas.items() if v]) 


def create_section4_1(wb, persons, activesections):
    sectionName = 'section4_1'
    ws = wb.create_sheet("Sección 4.1")
    title = "4.1 PARTICIPACIÓN EN GRUPOS DE INVESTIGACIÓN"
    columnTitles = [
        ("nombre del grupo", 40),
        ("programa", 30),
        ("lineas de investigación", 40),
        ("actividades", 80)

    ]
    createTitles(ws, columnTitles, title)

    fields = ['grupo', 'programa', 'linea_investiga', 'actividades']
    valueFromEnum = [['grupo', 'programa'],[enums.GruposInvesFacultad, enums.ProgramasFacultad]]
    specialCases = [['linea_investiga'],[specialCase4_1]]

    tables = getTables(sectionName)
    predata = db.session.query(tables[0]).all()

    data = []
    for dt in predata:
        lineasInv = [x for x in json.loads(dt.linea_investiga).values() if x]
        otherData = [getattr(dt, namefield) for namefield in fields if namefield != 'linea_investiga']
        if lineasInv or otherData:
            data.append(dt)

    extractorToExcel(data, ws, sectionName, persons, fields, wrap_text=['actividades'], valueFromEnum=valueFromEnum, specialCases=specialCases, activesections=activesections)
        
def create_section4_2(wb, persons):
    sectionName = 'section4_2'
    ws = wb.create_sheet("Sección 4.2")
    title = "4.2 REDES ACADÉMICAS DEL PROFESOR"
    columnTitles = [
        ("CVLAC link", 50),
        ("CVLAC ultima actualizacion", 20),
        ("Google Scolar link", 50),
        ("Google Scolar ultima actualizacion", 20),
        ("ResearchGate link", 50),
        ("ResearchGate ultima actualizacion", 20),
        ("ORcid link", 50),
        ("ORcid ultima actualizacion", 20)


    ]
    createTitles(ws, columnTitles, title)
    s4_1options = ['proceso', 'type_', 'observaciones_']

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    content = {}
    for entry in data:
        if entry.id_person not in content:
            content[entry.id_person] = {}
            for red in ['cvlac','google_scholar','researchgate','orcid']:
                content[entry.id_person][red] = {}
                content[entry.id_person][red]['status'] = ''
                content[entry.id_person][red]['link'] = 'No reporta'
                content[entry.id_person][red]['update'] = ''
        content[entry.id_person][entry.name_red]['status'] = entry.status
        content[entry.id_person][entry.name_red]['link'] = entry.link if entry.link else 'No reporta'
        content[entry.id_person][entry.name_red]['update'] = entry.update_date if entry.update_date else ''


    row = 2
    cells = []
    for id_person, data in content.items():
        row += 1
        col = 1
        cells.append(ws.cell(row=row, column=col, value=id_person))
        col += 1
        cells.append(ws.cell(row=row, column=col, value=persons[id_person]))
        for red in ['cvlac','google_scholar','researchgate','orcid']:
            col += 1
            cells.append(ws.cell(row=row, column=col, value=data[red]['link']))
            col += 1
            cells.append(ws.cell(row=row, column=col, value=data[red]['update']))
            cells[-1].alignment = Alignment(wrap_text=True)
    for cell in cells:
        cell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
        
def create_section4_3(wb, persons, activesections):
    sectionName = 'section4_3'
    ws = wb.create_sheet("Sección 4.3")
    title = "4.3 DESARROLLO DE PROYECTOS DE INVESTIGACIÓN O CREACIÓN"
    columnTitles = [
        ("nombre proyecto", 40),
        ("Código de radicado", 30),
        ("Tipo de convocatoria", 25),
        ("Nombre de la convocatoria", 40),
        ("¿El proyecto fue elegido?", 30),
        ("Entidad financiadora", 40),
        ("Tipo de participación", 30),
        ("participación de semillero", 30),
        ("Nombre del semillero", 50),
        ("estudiantes", 80),
        ("aporta a los ODS", 30),
        ("desarrollo de software para la creación artística", 30),
        ("I+D+I5", 30),

    ]
    createTitles(ws, columnTitles, title)
    
    fields = [
        'nombre_proyecto',
        'code',
        'tipo_convocatoria',
        'nombre_convocatoria',
        'elegido',
        'financiamiento',
        'tipo_participacion',
        'bool_semillero', 
        'semillero', 
        'estudiante',
        'ods', 
        'software',
        'idi5'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()
    itemsList=['estudiante']
    excludeofEmptyness=['estudiante']

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, itemsList=itemsList, wrap_text=['estudiante'], excludeofEmptyness=excludeofEmptyness)


def create_section4_4(wb, persons, activesections):
    sectionName = 'section4_4'
    ws = wb.create_sheet("Sección 4.4")
    title = "4.4 PRODUCTOS DE CREACIÓN"
    columnTitles = [
        ("Tipo de producto", 40),
        ("Nombre del producto", 30),
        ("Institución", 25),
        ("Alcance", 30),
        ("Fecha", 20),
        ("Link de difusión o nombre del catálogo", 40),
        ("Categoría de reconocimiento", 20)
    ]
    createTitles(ws, columnTitles, title)
    
    fields = [
        'tipo_producto',
        'nombre_producto',
        'institucion',
        'alcance',
        'fecha',
        'link',
        'categoria'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    translateResponse = [
        ["categoria"],
        {"tipo_a":"Tipo A", "tipo_b":"Tipo B", "no_responde":"No ha sido reconocido"}
    ]


    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, valueFromEnum=[['tipo_producto'],[enums.TiposProductoCreacion]], excludeofEmptyness=[], translateResponse=translateResponse)



def create_section4_5(wb, persons, activesections):
    sectionName = 'section4_5'
    ws = wb.create_sheet("Sección 4.5")
    title = "4.5 PUBLICACIÓN DE ARTÍCULOS"
    columnTitles = [
        ("titulo", 40),
        ("autores", 40),
        ("revista", 30),
        ("link", 40),
        ("Fecha", 20),
        ("issn", 40),
        ("index revista", 20)
    ]
    createTitles(ws, columnTitles, title)
    
    fields = [
        'titulo',
        'autores',
        'revista',
        'link',
        'fecha',
        'issn',
        'index_revista'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, itemsList=['autores'], excludeofEmptyness=['autores'])


def create_section4_6(wb, persons, activesections):
    sectionName = 'section4_6'
    ws = wb.create_sheet("Sección 4.6")
    title = "4.6 PUBLICACIÓN DE LIBROS O CAPITULOS DE LIBRO"
    columnTitles = [
        ("titulo", 40),
        ("autores", 40),
        ("editorial", 40),
        ("fecha", 20),
        ("issn", 40),
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'titulo',
        'autores',
        'editorial',
        'fecha',
        'issn'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, itemsList=['autores'], excludeofEmptyness=[])


def create_section4_7(wb, persons, activesections):
    sectionName = 'section4_7'
    ws = wb.create_sheet("Sección 4.7")
    title = "4.7 PARTICIPACIÓN EN EVENTOS ACADÉMICOS O ARTÍSTICOS NACIONALES E INTERNACIONALES"
    columnTitles = [
        ("nombre de la ponencia", 40),
        ("nombre del evento", 40),
        ("lugar", 40),
        ("fecha", 20),
        ("tipo participacion", 40),
        ("modalidad", 30),
        ("aprobado por consejo de facultad", 20),
        ("certificado", 40),
        ("participa semillero", 40),
        ("nombre semillero", 40),
        ("participantes", 40),
        ("movilidad", 40)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'titulo',
        'evento',
        'lugar',
        'fecha',
        'tipo_participacion',
        'modalidad',
        'aprobado',
        'name_file',
        'bool_semillero',
        'semillero',
        'participantes',
        'movilidad'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, itemsList=['participantes'], wrap_text=['participantes'], excludeofEmptyness=[])


def create_section4_8(wb, persons, activesections):
    sectionName = 'section4_8'
    ws = wb.create_sheet("Sección 4.8")
    title = "4.8 CONVENIOS INTERINSTITUCIONALES Y PRODUCTOS DERIVADOS DEL TRABAJO EN REDES"
    columnTitles = [
        ('convenio', 40),
        ('producto derivado', 40),
        ('grupo de investigación', 40),
        ('Impacto del proyecto en la formación de los estudiantes', 40)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'convenio',
        'producto',
        'grupo',
        'impacto'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections,wrap_text=['impacto'], valueFromEnum=[['grupo'],[enums.GruposInvesFacultad]], excludeofEmptyness=[])


def create_section4_9(wb, persons, activesections):
    sectionName = 'section4_9'
    ws = wb.create_sheet("Sección 4.9")
    title = "4.9 PATENTES O REGISTRO DE DERECHOS DE AUTOR"
    columnTitles = [
        ('tipo de producto', 40),
        ('registro', 40),
        ('fecha', 20)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'tipo_producto',
        'registro',
        'fecha'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, valueFromEnum=[['tipo_producto'],[enums.TipoProductosPatentesRegistroAutor]], excludeofEmptyness=[])


def create_section4_10(wb, persons, activesections):
    sectionName = 'section4_10'
    ws = wb.create_sheet("Sección 4.10")
    title = "4.10 DIRECCIÓN DE TESIS O TRABAJOS DE GRADO"
    columnTitles = [
        ('Nombre del trabajo de grado', 40),
        ('estudiantes', 40),
        ('nivel de formacion', 40),
        ('programa', 40),
        ('Estado actual del trabajo de grado', 40)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'titulo',
        'estudiante',
        'nivel_formacion',
        'programa',
        'status'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=['estudiante'], row=None, forcedValueFields=[], valueFromEnum=[['programa'],[enums.ProgramasFacultad]], specialCases=[], itemsList=['estudiante'], excludeofEmptyness=[])


def create_section4_11(wb, persons, activesections):
    sectionName = 'section4_11'
    ws = wb.create_sheet("Sección 4.11")
    title = "4.11 EVALUACIÓN DE TESIS O TRABAJOS DE GRADO"
    columnTitles = [
        ('Nombre del trabajo de grado ', 40),
        ('estudiantes', 40),
        ('nivel de formacion', 30),
        ('programa', 40)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'titulo',
        'estudiante',
        'nivel_formacion',
        'programa'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=['estudiante'], row=None, forcedValueFields=[], valueFromEnum=[['programa'],[enums.ProgramasFacultad]], specialCases=[], itemsList=['estudiante'], excludeofEmptyness=[])


def create_section4_12(wb, persons, activesections):
    sectionName = 'section4_12'
    ws = wb.create_sheet("Sección 4.12")
    title = "4.12 CREACIÓN, GESTIÓN Y COORDINACIÓN DE SEMILLEROS DE INVESTIGACIÓN"
    columnTitles = [
        ('Nombre del semillero', 40),
        ('total integrantes', 20),
        ('nuevos integrantes', 20),
        ('grupo de investigación', 40),
        ('programa', 40),
        ('actualizado en red sia', 20)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'semillero',
        'total_integrantes',
        'nuevos_integrantes',
        'grupo',
        'programa',
        'update_sia'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=[], row=None, forcedValueFields=[], valueFromEnum=[['grupo', 'programa'],[enums.GruposInvesFacultad, enums.ProgramasFacultad]], specialCases=[], itemsList=[], excludeofEmptyness=[])


def create_section4_13(wb, persons, activesections):
    sectionName = 'section4_13'
    ws = wb.create_sheet("Sección 4.13")
    title = "4.13 PARTICIPACIÓN EN CONVOCATORIAS DE SEMILLEROS DE INVESTIGACIÓN"
    columnTitles = [
        ('Nombre de la convocatoria', 40),
        ('tipo de convovatoria', 40),
        ('Nombre del proyecto', 40),
        ('semillero', 40),
        ('participantes', 40),
        ('fecha', 20),
        ('observaciones', 40)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'nombre',
        'tipo_convovatoria',
        'proyecto',
        'semillero',
        'participantes',
        'fecha',
        'observaciones'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=['observaciones', 'participantes'], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=['participantes'], excludeofEmptyness=[])


def create_section4_14(wb, persons, activesections):
    sectionName = 'section4_14'
    ws = wb.create_sheet("Sección 4.14")
    title = "4.14 MOVILIDAD REGIONAL, NACIONAL E INTERNACIONAL DE ESTUDIANTES Y SEMILLEROS DE INVESTIGACIÓN"
    columnTitles = [
        ('semillero', 40),
        ('Nombre del evento', 40),
        ('Nombre del proyecto', 40),
        ('lugar', 40),
        ('fecha', 40),
        ('estudiantes', 40)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'semillero',
        'evento',
        'proyecto',
        'lugar',
        'fecha',
        'estudiante'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=['estudiante'], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=['estudiante'], excludeofEmptyness=[])


def create_section4_15(wb, persons, activesections):
    sectionName = 'section4_15'
    ws = wb.create_sheet("Sección 4.15")
    title = "4.15 JÓVENES INVESTIGADORES APOYADOS PARA SU FORMACIÓN"
    columnTitles = [
        ('Nombre de la convocatoria', 40),
        ('estudiantes', 40),
        ('programa', 40)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'convocatoria',
        'estudiante',
        'programa'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=['estudiante'], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=['estudiante'], excludeofEmptyness=[])


def create_section4_16(wb, persons, activesections):
    sectionName = 'section4_16'
    ws = wb.create_sheet("Sección 4.16")
    title = "4.16 OTRAS ACTIVIDADES DE INVESTIGACIÓN, EXTENSIÓN O CREACIÓN"
    columnTitles = [
        ('Tipo de actividad', 40),
        ('Nombre del producto', 40),
        ('Institución', 40),
        ('Actividades llevadas a cabo', 40),
        ('Observaciones y/o detalles adicionales', 40)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'tipo_actividad',
        'nombre',
        'institucion',
        'actividades',
        'detalles'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=['actividades', 'detalles'], row=None, forcedValueFields=[], valueFromEnum=[['tipo_actividad'],[enums.OtrasInvestigacionTipoParticipacion]], specialCases=[], itemsList=[], excludeofEmptyness=[])


def create_section4_17(wb, persons, activesections):
    sectionName = 'section4_17'
    ws = wb.create_sheet("Sección 4.17")
    title = "4.17 OTRAS ACTIVIDADES RELACIONADAS CON RECURSOS, ESTRATEGIAS Y CAPACITACIONES"
    columnTitles = [
        ('tipo_actividad', 40),
        ('nombre de la actividad', 40),
        ('programa', 40),
        ('asignatura', 40),
        ('fecha', 40),
        ('finalizado', 40),
        ('alcance', 40),
        ('impacto', 40),
        ('detalles y/o observaciones', 40),
        ('ofertado por', 40)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'tipo_actividad',
        'nombre',
        'programa',
        'asignatura',
        'fecha',
        'finalizado',
        'alcance',
        'impacto',
        'detalles',
        'ofertado'
    ]

    tiposPart = {
        "creacion":["nombre", "programa", "asignatura", "fecha", "impacto", "detalles"],
        "design": ["nombre", "programa", "asignatura", "fecha", "impacto", "detalles"],
        "actividades": ["nombre", "alcance", "impacto", "detalles"],
        "participacion": ["nombre", "ofertado", "finalizado", "impacto", "detalles"],
        "capacitacion": ["nombre", "ofertado", "finalizado", "impacto", "detalles"]
        }

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    filteredData = {key:{} for key in tiposPart.keys()}
    wrap_text = [
        'impacto',
        'detalles',
    ]

    row = 2
    cells = []
    for entry in data:
        if activesections and activesections[entry.id_person]['section4_17'] != 'si':
            continue
        if all([getattr(entry, namefield) is None for namefield in fields]):
            continue
        row += 1
        cells.append(ws.cell(row=row, column=1, value=entry.id_person))
        cells.append(ws.cell(row=row, column=2, value=persons[entry.id_person]))
        for col, namefield in enumerate(fields, 3):
            if namefield == 'tipo_actividad' or not getattr(entry, 'tipo_actividad'):
                value = enums.otrasActividades4_17[getattr(entry, namefield)].value if getattr(entry, namefield) else ''
            else:
                if namefield in tiposPart[entry.tipo_actividad]:
                    value = getattr(entry, namefield) if getattr(entry, namefield) else ''
                else:
                    value = 'n.a.'
            cells.append(ws.cell(row=row, column=col, value=value))
            if namefield in wrap_text:
                cells[-1].alignment = Alignment(wrap_text=True)

    for cell in cells:
        cell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)

def create_section5_1(wb, persons, activesections):
    sectionName = 'section5_1'
    ws = wb.create_sheet("Sección 5.1")
    title = "5.1 PROPUESTAS DE EDUCACIÓN CONTÍNUA PRESENTADAS"
    columnTitles = [
        ('Nombre de la propuesta', 40),
        ('Tipo de propuesta', 40),
        ('Fecha de inicio', 40),
        ('Fecha de finalización', 40),
        ('Modalidad', 40),
        ('componente internacional', 40),
        ('Atiende a población diferencial', 40),
        ('población diferencial', 40),
        ('convenio con otra entidad', 40),
        ('Nombre de la entidad', 40)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'nombre',
        'tipo_propuesta',
        'fecha_inicio',
        'fecha_fin',
        'modalidad',
        'internacional',
        'bool_diferencial',
        'diferencial',
        'bool_entidad',
        'entidad'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=[], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=[], excludeofEmptyness=[])

def create_section5_2(wb, persons, activesections):
    sectionName = 'section5_2'
    ws = wb.create_sheet("Sección 5.2")
    title = "5.2 GESTIÓN DE CONVENIOS O ALIANZAS PARA ACCESO A LA CULTURA"
    columnTitles = [
        ('entidad', 40),
        ('actividades', 40),
        ('tipo_vinculo', 40),
        ('beneficiarios', 40)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'entidad',
        'actividades',
        'tipo_vinculo',
        'beneficiarios'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=[], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=[], excludeofEmptyness=[])

def create_section5_3(wb, persons, activesections):
    sectionName = 'section5_3'
    ws = wb.create_sheet("Sección 5.3")
    title = "5.3 PROYECTOS DE PROYECCIÓN SOCIAL EJECUTADOS CON INTERVENCIÓN EN POBLACIÓN PRIORIZADA, CON ENFOQUE DIFERENCIAL O CAPACIDADES DIVERSAS"
    columnTitles = [
        ('Nombre del proyecto', 40),
        ('Fecha de inicio', 40),
        ('Población beneficiada', 40),
        ('participantes', 40),
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'nombre_proyecto',
        'fecha_inicio',
        'beneficiarios',
        'participantes'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=['participantes'], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=['participantes'], excludeofEmptyness=[])

def create_section6_1(wb, persons):
    sectionName = 'section6_1'
    ws = wb.create_sheet("Sección 6.1")
    title = "6.1 ASIGNACIÓN MISIONAL DE BIENESTAR SEGÚN PTA"
    columnTitles = [
        ('representante misional', 20),
        ('tutor rendimiento', 20),
        ('tutor SAT', 20),
        ('coordinador', 20)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'representante_misional',
        'tutor_rendimiento',
        'tutor_SAT',
        'coordinador'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=None, wrap_text=[], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=[], excludeofEmptyness=[])

def create_section6_2(wb, persons, activesections):
    sectionName = 'section6_2'
    ws = wb.create_sheet("Sección 6.2")
    title = "6.2 TUTORÍAS Y ESTUDIANTES ATENDIDOS DESDE BIENESTAR - BAJO RENDIMIENTO"
    columnTitles = [
        ('soporte', 50)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'name_file'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=[], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=[], excludeofEmptyness=[])

def create_section6_3(wb, persons, activesections):
    sectionName = 'section6_3'
    ws = wb.create_sheet("Sección 6.3")
    title = "6.3 TUTORÍAS Y ESTUDIANTES ATENDIDOS DESDE BIENESTAR - TUTORIAS SAT"
    columnTitles = [
        ('soporte', 50)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'name_file'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=[], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=[], excludeofEmptyness=[])

def create_section6_4(wb, persons, activesections):
    sectionName = 'section6_4'
    ws = wb.create_sheet("Sección 6.4")
    title = "6.4 TUTORÍAS Y ESTUDIANTES ATENDIDOS DESDE BIENESTAR - MATERIA UNICA"
    columnTitles = [
        ('soporte', 50)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'name_file'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=[], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=[], excludeofEmptyness=[])

def create_section6_7(wb, persons, activesections):
    sectionName = 'section6_7'
    ws = wb.create_sheet("Sección 6.5 (section6_7)")
    title = "6.5 TUTORÍAS Y ESTUDIANTES ATENDIDOS DESDE BIENESTAR - CASOS ESPECIALES"
    columnTitles = [
        ('soporte', 50)
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'name_file'
    ]

    tables = getTables(sectionName)
    data = db.session.query(tables[0]).all()

    extractorToExcel(data, ws, sectionName, persons, fields, activesections=activesections, wrap_text=[], row=None, forcedValueFields=[], valueFromEnum=[], specialCases=[], itemsList=[], excludeofEmptyness=[])

def create_activateSections(wb, persons, activesections):
    '''
    por petición del cliente, la pregunta 6.7 pasó a ser la 6.5 y no se usaron la 6.5 y 6.6
    Sin embargo, como el cambio puede que sea temporal, la base de datos mantiene los nombres originales
    '''
    sectionName = 'section6_4'
    ws = wb.create_sheet("Secciones Activas")
    title = "CONTROL DE SECCIONES ACTIVAS POR USUARIO"
    columnTitles = [
        ('section4_1', 14), 
        ('section4_3', 14),
        ('section4_4', 14),
        ('section4_5', 14), 
        ('section4_6', 14),
        ('section4_7', 14),
        ('section4_8', 14), 
        ('section4_9', 14), 
        ('section4_10', 14), 
        ('section4_11', 14), 
        ('section4_12', 14), 
        ('section4_13', 14), 
        ('section4_14', 14), 
        ('section4_15', 14), 
        ('section4_16', 14),
        ('section4_17', 14),
        ('section5_1', 14),
        ('section5_2', 14),
        ('section5_3', 14),
        ('section6_2', 14),
        ('section6_3', 14), 
        ('section6_4', 14), 
        ('section6_5', 14),
    ]

    createTitles(ws, columnTitles, title)
    
    fields = [
        'section4_1',
        'section4_3',
        'section4_4',
        'section4_5',
        'section4_6',
        'section4_7',
        'section4_8',
        'section4_9',
        'section4_10',
        'section4_11',
        'section4_12',
        'section4_13',
        'section4_14',
        'section4_15',
        'section4_16',
        'section4_17',
        'section5_1',
        'section5_2',
        'section5_3',
        'section6_2',
        'section6_3',
        'section6_4',
        'section6_7'
    ]
    
    row = 2
    cells = []
    for id_person, entry in activesections.items():
        row += 1
        col = 1
        cells.append(ws.cell(row=row, column=col, value=id_person))
        col += 1
        cells.append(ws.cell(row=row, column=col, value=persons[id_person]))
        for namefield in fields:
            col += 1
            cells.append(ws.cell(row=row, column=col, value=entry[namefield]))

    for cell in cells:
        cell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
